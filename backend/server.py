from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import uuid
import sqlite3
import json
import logging
import io
import csv
import smtplib
import bcrypt
import jwt
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, Query, Form
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field, EmailStr
from apscheduler.schedulers.background import BackgroundScheduler


# ---- Config ----
# In production Railway sets DATA_DIR to a mounted volume path
DATA_DIR = Path(os.environ.get("DATA_DIR", str(ROOT_DIR)))
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / os.environ.get("DB_FILE", "schoolassets.db")
JWT_SECRET = os.environ.get("JWT_SECRET", "dev-secret-change-me")
JWT_ALGO = "HS256"
UPLOADS_DIR = DATA_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

# Email config (set in Railway environment variables)
SMTP_HOST     = os.environ.get("SMTP_HOST", "")
SMTP_PORT     = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER     = os.environ.get("SMTP_USER", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
SMTP_FROM     = os.environ.get("SMTP_FROM", SMTP_USER)
APP_URL       = os.environ.get("APP_URL", "https://your-app.railway.app")

# Loan notification recipients (set in Railway environment variables)
# General inbox that receives new-request and return-complete alerts
LOAN_NOTIFY_EMAIL  = os.environ.get("LOAN_NOTIFY_EMAIL", "")
# Campus heads — keyed by campus slug (KK = Kuala Kencana, TB = Tembagapura)
HEAD_KK_EMAIL       = os.environ.get("HEAD_KK_EMAIL", "")        # Head of Campus – YPJ Kuala Kencana
HEAD_TB_EMAIL       = os.environ.get("HEAD_TB_EMAIL", "")        # Head of Campus – YPJ Tembagapura
HEAD_ADMIN_KK_EMAIL = os.environ.get("HEAD_ADMIN_KK_EMAIL", "")  # Head of Administration – YPJ Kuala Kencana
HEAD_ADMIN_TB_EMAIL = os.environ.get("HEAD_ADMIN_TB_EMAIL", "")  # Head of Administration – YPJ Tembagapura

# Map campus name → (head-of-campus email, head-of-admin email)
CAMPUS_HEAD_EMAILS: dict = {
    "YPJ Kuala Kencana": [HEAD_KK_EMAIL, HEAD_ADMIN_KK_EMAIL],
    "YPJ Tembagapura":   [HEAD_TB_EMAIL, HEAD_ADMIN_TB_EMAIL],
}

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = FastAPI(title="School Assets Management API")
api = APIRouter(prefix="/api")


# ---- Database ----
def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    conn = get_conn()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id TEXT PRIMARY KEY,
        email TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'staff',
        department TEXT,
        campus TEXT NOT NULL DEFAULT '',
        password_hash TEXT NOT NULL,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS assets (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        asset_tag TEXT NOT NULL,
        category TEXT NOT NULL,
        description TEXT DEFAULT '',
        campus TEXT NOT NULL DEFAULT '',
        location TEXT NOT NULL,
        status TEXT NOT NULL DEFAULT 'active',
        purchase_price REAL NOT NULL,
        purchase_date TEXT NOT NULL,
        useful_life_years INTEGER NOT NULL DEFAULT 5,
        warranty_end_date TEXT,
        serial_number TEXT DEFAULT '',
        supplier TEXT DEFAULT '',
        assigned_to_user_id TEXT,
        assigned_to_name TEXT,
        photo_path TEXT,
        documents TEXT DEFAULT '[]',
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        created_by TEXT
    );

    CREATE TABLE IF NOT EXISTS ownership_logs (
        id TEXT PRIMARY KEY,
        asset_id TEXT NOT NULL,
        from_name TEXT,
        from_user_id TEXT,
        to_name TEXT,
        to_user_id TEXT,
        note TEXT DEFAULT '',
        by TEXT NOT NULL,
        by_name TEXT NOT NULL,
        at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS faults (
        id TEXT PRIMARY KEY,
        asset_id TEXT NOT NULL,
        title TEXT NOT NULL,
        description TEXT NOT NULL,
        severity TEXT NOT NULL DEFAULT 'medium',
        status TEXT NOT NULL DEFAULT 'open',
        reported_by TEXT NOT NULL,
        reported_by_name TEXT NOT NULL,
        resolution_note TEXT,
        resolved_at TEXT,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS compliance (
        id TEXT PRIMARY KEY,
        asset_id TEXT,
        title TEXT NOT NULL,
        description TEXT DEFAULT '',
        category TEXT NOT NULL,
        due_date TEXT NOT NULL,
        frequency TEXT NOT NULL DEFAULT 'annual',
        status TEXT NOT NULL DEFAULT 'pending',
        completed_date TEXT,
        note TEXT,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS files (
        id TEXT PRIMARY KEY,
        storage_path TEXT NOT NULL,
        original_filename TEXT NOT NULL,
        content_type TEXT NOT NULL,
        size INTEGER NOT NULL,
        uploaded_by TEXT NOT NULL,
        is_deleted INTEGER NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS loan_requests (
        id TEXT PRIMARY KEY,
        asset_id TEXT NOT NULL,
        asset_name TEXT NOT NULL,
        asset_tag TEXT NOT NULL,
        requested_by TEXT NOT NULL,
        requested_by_name TEXT NOT NULL,
        purpose TEXT DEFAULT '',
        start_date TEXT,
        end_date TEXT,
        status TEXT NOT NULL DEFAULT 'pending',
        reviewed_by TEXT,
        reviewed_by_name TEXT,
        reviewed_at TEXT,
        reject_reason TEXT,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS transfer_requests (
        id TEXT PRIMARY KEY,
        asset_id TEXT NOT NULL,
        asset_name TEXT NOT NULL,
        asset_tag TEXT NOT NULL,
        from_name TEXT,
        to_name TEXT NOT NULL,
        note TEXT DEFAULT '',
        requested_by TEXT NOT NULL,
        requested_by_name TEXT NOT NULL,
        status TEXT NOT NULL DEFAULT 'pending',
        reviewed_by TEXT,
        reviewed_by_name TEXT,
        reviewed_at TEXT,
        reject_reason TEXT,
        created_at TEXT NOT NULL
    );
    """)
    conn.commit()
    # Migrations for existing databases
    cols = [r[1] for r in conn.execute("PRAGMA table_info(assets)").fetchall()]
    if "campus" not in cols:
        conn.execute("ALTER TABLE assets ADD COLUMN campus TEXT NOT NULL DEFAULT ''")
        conn.commit()
    if "asset_type" not in cols:
        conn.execute("ALTER TABLE assets ADD COLUMN asset_type TEXT NOT NULL DEFAULT 'School Asset'")
        conn.commit()
    user_cols = [r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
    if "campus" not in user_cols:
        conn.execute("ALTER TABLE users ADD COLUMN campus TEXT NOT NULL DEFAULT ''")
        conn.commit()
    if "supervisor" not in user_cols:
        conn.execute("ALTER TABLE users ADD COLUMN supervisor TEXT DEFAULT ''")
        conn.commit()
    if "employee_id" not in user_cols:
        conn.execute("ALTER TABLE users ADD COLUMN employee_id TEXT DEFAULT ''")
        conn.commit()
    if "supervisor_name" not in user_cols:
        conn.execute("ALTER TABLE users ADD COLUMN supervisor_name TEXT DEFAULT ''")
        conn.commit()
    # Loan request workflow columns (deliver / return / reminder)
    loan_cols = [r[1] for r in conn.execute("PRAGMA table_info(loan_requests)").fetchall()]
    for col, defn in [
        ("pre_condition",      "TEXT DEFAULT ''"),
        ("pre_condition_note", "TEXT DEFAULT ''"),
        ("pre_image_path",     "TEXT DEFAULT ''"),
        ("post_condition",     "TEXT DEFAULT ''"),
        ("post_condition_note","TEXT DEFAULT ''"),
        ("post_image_path",    "TEXT DEFAULT ''"),
        ("delivered_at",       "TEXT"),
        ("returned_at",        "TEXT"),
    ]:
        if col not in loan_cols:
            conn.execute(f"ALTER TABLE loan_requests ADD COLUMN {col} {defn}")
    conn.commit()
    conn.close()


def row_to_dict(row) -> dict:
    if row is None:
        return None
    d = dict(row)
    if "documents" in d and isinstance(d["documents"], str):
        try:
            d["documents"] = json.loads(d["documents"])
        except Exception:
            d["documents"] = []
    return d


# ---- Auth helpers ----
def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id, "email": email, "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")
    conn = get_conn()
    row = conn.execute("SELECT * FROM users WHERE id = ?", (payload["sub"],)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(401, "User not found")
    user = row_to_dict(row)
    user.pop("password_hash", None)
    return user


async def admin_required(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin access required")
    return user


# ---- Models ----
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Literal["admin", "staff"] = "staff"
    department: Optional[str] = None
    campus: Optional[str] = ""
    supervisor: Optional[str] = ""
    supervisor_name: Optional[str] = ""
    employee_id: Optional[str] = ""


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class AssetIn(BaseModel):
    name: str
    asset_tag: str
    category: str
    asset_type: Optional[str] = "School Asset"
    description: Optional[str] = ""
    campus: Optional[str] = ""
    location: str
    status: Literal["active", "in_repair", "retired", "lost"] = "active"
    purchase_price: float
    purchase_date: str
    useful_life_years: int = 5
    warranty_end_date: Optional[str] = None
    serial_number: Optional[str] = ""
    supplier: Optional[str] = ""
    assigned_to_user_id: Optional[str] = None
    assigned_to_name: Optional[str] = None
    photo_path: Optional[str] = None
    documents: List[dict] = []


class AssetUpdate(BaseModel):
    name: Optional[str] = None
    asset_tag: Optional[str] = None
    category: Optional[str] = None
    asset_type: Optional[str] = None
    description: Optional[str] = None
    campus: Optional[str] = None
    location: Optional[str] = None
    status: Optional[str] = None
    purchase_price: Optional[float] = None
    purchase_date: Optional[str] = None
    useful_life_years: Optional[int] = None
    warranty_end_date: Optional[str] = None
    serial_number: Optional[str] = None
    supplier: Optional[str] = None
    photo_path: Optional[str] = None
    documents: Optional[List[dict]] = None


class TransferIn(BaseModel):
    new_owner_user_id: Optional[str] = None
    new_owner_name: str
    note: Optional[str] = ""


class FaultIn(BaseModel):
    asset_id: str
    title: str
    description: str
    severity: Literal["low", "medium", "high"] = "medium"


class FaultUpdate(BaseModel):
    status: Optional[Literal["open", "in_progress", "resolved"]] = None
    resolution_note: Optional[str] = None


class ComplianceIn(BaseModel):
    asset_id: Optional[str] = None
    title: str
    description: Optional[str] = ""
    category: str
    due_date: str
    frequency: Literal["once", "monthly", "quarterly", "annual"] = "annual"


class ComplianceUpdate(BaseModel):
    status: Optional[Literal["pending", "completed", "overdue"]] = None
    completed_date: Optional[str] = None
    note: Optional[str] = None
    due_date: Optional[str] = None


# ---- Utility ----
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def calc_depreciation(purchase_price: float, purchase_date: str, useful_life_years: int) -> dict:
    try:
        pdate = datetime.fromisoformat(purchase_date.replace("Z", "+00:00"))
        if pdate.tzinfo is None:
            pdate = pdate.replace(tzinfo=timezone.utc)
    except Exception:
        return {"current_value": purchase_price, "accumulated": 0, "annual": 0, "age_years": 0}
    age = (datetime.now(timezone.utc) - pdate).days / 365.25
    annual = purchase_price / max(useful_life_years, 1)
    accumulated = min(annual * age, purchase_price)
    current = max(purchase_price - accumulated, 0)
    return {
        "current_value": round(current, 2),
        "accumulated": round(accumulated, 2),
        "annual": round(annual, 2),
        "age_years": round(age, 2),
    }


# ---- Email helper ----
def send_email(to_addr: str, subject: str, body_html: str, cc: list = None):
    """Send an HTML email; cc is an optional list of additional recipients."""
    if not SMTP_HOST or not SMTP_USER:
        logger.warning("Email not configured — skipping send to %s", to_addr)
        return
    try:
        clean_cc = [a for a in (cc or []) if a and a != to_addr]
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = SMTP_FROM or SMTP_USER
        msg["To"]      = to_addr
        if clean_cc:
            msg["Cc"] = ", ".join(clean_cc)
        msg.attach(MIMEText(body_html, "html"))
        all_recipients = [to_addr] + clean_cc
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.ehlo()
            s.starttls()
            s.login(SMTP_USER, SMTP_PASSWORD)
            s.sendmail(SMTP_FROM or SMTP_USER, all_recipients, msg.as_string())
        logger.info("Email sent to %s (CC: %s) — %s", to_addr, clean_cc, subject)
    except Exception as exc:
        logger.error("Email send failed: %s", exc)


def get_loan_cc(req: dict, conn, extra: list = None) -> list:
    """
    Build the CC list for a loan-related email.

    Includes:
      - The borrower's supervisor (stored in users.supervisor)
      - Head of campus and head of administration for the asset's campus
      - Any extra addresses (e.g. LOAN_NOTIFY_EMAIL for new-request / return alerts)
    """
    cc = list(extra or [])

    # Supervisor of the borrower
    u = conn.execute("SELECT supervisor FROM users WHERE id=?", (req["requested_by"],)).fetchone()
    if u and u["supervisor"]:
        cc.append(u["supervisor"])

    # Campus heads based on the asset's campus
    asset = conn.execute("SELECT campus FROM assets WHERE id=?", (req["asset_id"],)).fetchone()
    campus = (asset["campus"] if asset else "") or ""
    for addr in CAMPUS_HEAD_EMAILS.get(campus, []):
        if addr:
            cc.append(addr)

    # Deduplicate, preserve order, drop blanks
    seen: set = set()
    result = []
    for a in cc:
        if a and a not in seen:
            seen.add(a)
            result.append(a)
    return result


def _loan_html(req: dict, heading: str, detail_lines: list, footer: str = "") -> str:
    """Generic HTML body for loan notification emails."""
    rows = "".join(
        f"<tr><td style='padding:4px 8px;color:#64748b;'>{k}</td>"
        f"<td style='padding:4px 8px;font-weight:500;'>{v}</td></tr>"
        for k, v in detail_lines
    )
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
      <div style="background:#0f172a;color:#fff;padding:16px 24px;">
        <strong>YPJ School Assets Management</strong>
      </div>
      <div style="padding:24px;">
        <h2 style="margin-top:0;">{heading}</h2>
        <table style="border-collapse:collapse;width:100%;margin-bottom:16px;">
          {rows}
        </table>
        {footer}
        <p style="margin-top:24px;">
          <a href="{APP_URL}" style="background:#0f172a;color:#fff;padding:10px 20px;text-decoration:none;">
            Open Asset Registry
          </a>
        </p>
      </div>
      <div style="background:#f8fafc;padding:12px 24px;font-size:12px;color:#94a3b8;">
        YPJ School Assets Management &mdash; automated notification
      </div>
    </div>
    """


# ---- Scheduler: return-date reminders ----
def send_return_reminders():
    """Runs daily at 08:00; emails borrowers whose return date is tomorrow, with CC chain."""
    tomorrow = (datetime.now(timezone.utc) + timedelta(days=1)).date().isoformat()
    conn = get_conn()
    rows = conn.execute(
        """SELECT lr.*, u.email AS borrower_email
           FROM loan_requests lr
           LEFT JOIN users u ON u.id = lr.requested_by
           WHERE lr.status = 'active'
             AND substr(lr.end_date, 1, 10) = ?""",
        (tomorrow,),
    ).fetchall()
    for r in rows:
        req = dict(r)
        email = req.get("borrower_email", "")
        if not email:
            continue
        asset_name = req["asset_name"]
        subject = f'[YPJ Assets] Reminder: Please return "{asset_name}" tomorrow'
        body = _loan_html(req, "Return Reminder", [
            ("Asset",     f"{req['asset_name']} ({req['asset_tag']})"),
            ("Borrower",  req["requested_by_name"]),
            ("Return by", tomorrow),
        ], "<p>Please return the asset to the admin before the end of the day.</p>")
        cc = get_loan_cc(req, conn, extra=[LOAN_NOTIFY_EMAIL] if LOAN_NOTIFY_EMAIL else [])
        send_email(email, subject, body, cc=cc)
    conn.close()
    logger.info("Return reminders checked — %d sent", len(rows))


# ---- Auth endpoints ----
@api.post("/auth/register")
async def register(body: UserCreate, response: Response):
    conn = get_conn()
    email = body.email.lower()
    existing = conn.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
    if existing:
        conn.close()
        raise HTTPException(400, "Email already registered")
    uid = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO users (id,email,name,role,department,campus,supervisor,supervisor_name,employee_id,password_hash,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (uid, email, body.name, body.role, body.department, body.campus or "", body.supervisor or "", body.supervisor_name or "", body.employee_id or "", hash_password(body.password), now_iso()),
    )
    conn.commit()
    conn.close()
    token = create_token(uid, email, body.role)
    response.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=7*24*3600, path="/")
    return {"id": uid, "email": email, "name": body.name, "role": body.role, "department": body.department, "token": token}


@api.post("/auth/login")
async def login(body: LoginIn, response: Response):
    conn = get_conn()
    email = body.email.lower()
    row = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
    conn.close()
    if not row or not verify_password(body.password, row["password_hash"]):
        raise HTTPException(401, "Invalid email or password")
    user = row_to_dict(row)
    token = create_token(user["id"], user["email"], user["role"])
    response.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=7*24*3600, path="/")
    return {
        "id": user["id"], "email": user["email"], "name": user["name"],
        "role": user["role"], "department": user.get("department"),
        "campus": user.get("campus"),
        "token": token,
    }


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


# ---- Users ----
@api.get("/users")
async def list_users(user: dict = Depends(get_current_user)):
    conn = get_conn()
    rows = conn.execute("SELECT id,email,name,role,department,campus,supervisor,supervisor_name,employee_id,created_at FROM users ORDER BY created_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@api.post("/users")
async def create_user(body: UserCreate, _: dict = Depends(admin_required)):
    return await register(body, Response())


@api.delete("/users/{uid}")
async def delete_user(uid: str, _: dict = Depends(admin_required)):
    conn = get_conn()
    res = conn.execute("DELETE FROM users WHERE id = ?", (uid,))
    conn.commit()
    conn.close()
    if res.rowcount == 0:
        raise HTTPException(404, "User not found")
    return {"ok": True}


class UserPatch(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None
    department: Optional[str] = None
    campus: Optional[str] = None
    supervisor: Optional[str] = None
    supervisor_name: Optional[str] = None
    employee_id: Optional[str] = None
    password: Optional[str] = None


@api.put("/users/{uid}")
async def update_user(uid: str, body: UserPatch, _: dict = Depends(admin_required)):
    updates = {k: v for k, v in body.model_dump(exclude_none=True).items() if v is not None}
    if not updates:
        raise HTTPException(400, "No fields to update")
    if "password" in updates:
        updates["password_hash"] = hash_password(updates.pop("password"))
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    params = list(updates.values()) + [uid]
    conn = get_conn()
    res = conn.execute(f"UPDATE users SET {set_clause} WHERE id = ?", params)
    if res.rowcount == 0:
        conn.close()
        raise HTTPException(404, "User not found")
    conn.commit()
    row = conn.execute("SELECT id,email,name,role,department,campus,supervisor,supervisor_name,employee_id,created_at FROM users WHERE id = ?", (uid,)).fetchone()
    conn.close()
    return dict(row)


# ---- File upload ----
@api.post("/upload")
async def upload_file(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "bin"
    file_id = str(uuid.uuid4())
    rel_path = f"{user['id']}/{file_id}.{ext}"
    dest = UPLOADS_DIR / user['id']
    dest.mkdir(exist_ok=True)
    full_path = dest / f"{file_id}.{ext}"
    data = await file.read()
    full_path.write_bytes(data)
    content_type = file.content_type or "application/octet-stream"
    conn = get_conn()
    conn.execute(
        "INSERT INTO files (id,storage_path,original_filename,content_type,size,uploaded_by,is_deleted,created_at) VALUES (?,?,?,?,?,?,0,?)",
        (file_id, rel_path, file.filename, content_type, len(data), user["id"], now_iso()),
    )
    conn.commit()
    conn.close()
    return {"path": rel_path, "name": file.filename, "size": len(data), "content_type": content_type}


@api.get("/files/{path:path}")
async def download_file(path: str, request: Request, auth: Optional[str] = Query(None)):
    token = request.cookies.get("access_token")
    if not token:
        ah = request.headers.get("Authorization", "")
        if ah.startswith("Bearer "):
            token = ah[7:]
    if not token and auth:
        token = auth
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except Exception:
        raise HTTPException(401, "Invalid token")
    full_path = UPLOADS_DIR / path
    if not full_path.exists():
        raise HTTPException(404, "File not found")
    conn = get_conn()
    row = conn.execute("SELECT * FROM files WHERE storage_path = ? AND is_deleted = 0", (path,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "File not found")
    return FileResponse(str(full_path), media_type=row["content_type"], filename=row["original_filename"])


# ---- Assets ----
@api.get("/assets")
async def list_assets(
    user: dict = Depends(get_current_user),
    search: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    location: Optional[str] = None,
    campus: Optional[str] = None,
    assigned_to: Optional[str] = None,
):
    conn = get_conn()
    conditions = []
    params = []
    if category:
        conditions.append("category = ?")
        params.append(category)
    if status:
        conditions.append("status = ?")
        params.append(status)
    if campus:
        conditions.append("campus = ?")
        params.append(campus)
    if location:
        conditions.append("location = ?")
        params.append(location)
    if assigned_to:
        conditions.append("assigned_to_user_id = ?")
        params.append(assigned_to)
    if search:
        conditions.append("(name LIKE ? OR asset_tag LIKE ? OR serial_number LIKE ?)")
        like = f"%{search}%"
        params.extend([like, like, like])
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    rows = conn.execute(f"SELECT * FROM assets {where} ORDER BY created_at DESC", params).fetchall()
    conn.close()
    assets = [row_to_dict(r) for r in rows]
    for a in assets:
        a["depreciation"] = calc_depreciation(a["purchase_price"], a["purchase_date"], a.get("useful_life_years", 5))
    return assets


@api.post("/assets")
async def create_asset(body: AssetIn, user: dict = Depends(get_current_user)):
    aid = str(uuid.uuid4())
    ts = now_iso()
    conn = get_conn()
    conn.execute("""
        INSERT INTO assets (id,name,asset_tag,category,asset_type,description,campus,location,status,
            purchase_price,purchase_date,useful_life_years,warranty_end_date,
            serial_number,supplier,assigned_to_user_id,assigned_to_name,
            photo_path,documents,created_at,updated_at,created_by)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        aid, body.name, body.asset_tag, body.category, body.asset_type or "School Asset",
        body.description or "", body.campus or "", body.location, body.status,
        body.purchase_price, body.purchase_date, body.useful_life_years, body.warranty_end_date,
        body.serial_number or "", body.supplier or "", body.assigned_to_user_id, body.assigned_to_name,
        body.photo_path, json.dumps(body.documents), ts, ts, user["id"],
    ))
    if body.assigned_to_name:
        conn.execute("""
            INSERT INTO ownership_logs (id,asset_id,from_name,from_user_id,to_name,to_user_id,note,by,by_name,at)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (str(uuid.uuid4()), aid, None, None, body.assigned_to_name,
              body.assigned_to_user_id, "Initial assignment", user["id"], user["name"], ts))
    conn.commit()
    row = conn.execute("SELECT * FROM assets WHERE id = ?", (aid,)).fetchone()
    conn.close()
    doc = row_to_dict(row)
    doc["depreciation"] = calc_depreciation(doc["purchase_price"], doc["purchase_date"], doc.get("useful_life_years", 5))
    return doc


IMPORT_COLUMNS = [
    "name", "asset_tag", "category", "asset_type", "description",
    "campus", "location", "status", "purchase_price", "purchase_date",
    "useful_life_years", "warranty_end_date", "serial_number", "supplier",
    "assigned_to_name",
]

VALID_CATEGORIES = [
    "IT Equipment", "AV Equipment", "Furniture", "Lab Equipment",
    "Vehicles", "Books", "Sports", "Music Equipment", "Other",
]
VALID_STATUSES = ["active", "in_repair", "retired", "lost"]
VALID_CAMPUSES = ["YPJ Kuala Kencana", "YPJ Tembagapura", ""]


def _validate_import_row(row: dict, row_num: int, seen_tags: set) -> list:
    """Return a list of error strings for this row (empty = valid)."""
    errors = []
    name = (row.get("name") or "").strip()
    tag  = (row.get("asset_tag") or "").strip()
    cat  = (row.get("category") or "").strip()
    loc  = (row.get("location") or "").strip()
    status = (row.get("status") or "active").strip()
    campus = (row.get("campus") or "").strip()

    if not name:
        errors.append("name is required")
    if not tag:
        errors.append("asset_tag is required")
    elif tag in seen_tags:
        errors.append(f'duplicate asset_tag "{tag}" within file')
    if not cat:
        errors.append("category is required")
    elif cat not in VALID_CATEGORIES:
        errors.append(f'unknown category "{cat}" — must be one of: {", ".join(VALID_CATEGORIES)}')
    if not loc:
        errors.append("location is required")
    if status not in VALID_STATUSES:
        errors.append(f'invalid status "{status}" — must be: active, in_repair, retired, lost')
    if campus and campus not in VALID_CAMPUSES:
        errors.append(f'unknown campus "{campus}"')

    price_str = (row.get("purchase_price") or "").strip()
    if price_str:
        try:
            float(price_str)
        except ValueError:
            errors.append(f'purchase_price "{price_str}" is not a number')

    life_str = (row.get("useful_life_years") or "").strip()
    if life_str:
        try:
            int(life_str)
        except ValueError:
            errors.append(f'useful_life_years "{life_str}" must be an integer')

    for date_col in ("purchase_date", "warranty_end_date"):
        val = (row.get(date_col) or "").strip()
        if val:
            try:
                datetime.strptime(val, "%Y-%m-%d")
            except ValueError:
                errors.append(f'{date_col} "{val}" must be YYYY-MM-DD')

    return errors


def _parse_import_file(content: bytes, filename: str) -> list:
    """
    Parse CSV or XLSX bytes. Returns list of dicts, each with the IMPORT_COLUMNS keys.
    Raises HTTPException 400 for unparseable files.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = [str(c).strip() if c else "" for c in next(rows_iter, [])]
            rows = []
            for r in rows_iter:
                rows.append({header[i]: (str(r[i]).strip() if r[i] is not None else "") for i in range(len(header))})
        except Exception as exc:
            raise HTTPException(400, f"Could not parse Excel file: {exc}")
    else:
        try:
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = [dict(r) for r in reader]
        except Exception as exc:
            raise HTTPException(400, f"Could not parse CSV file: {exc}")
    return rows


@api.get("/assets/import-template")
async def download_import_template(_: dict = Depends(admin_required)):
    header = ",".join(IMPORT_COLUMNS)
    example = (
        'Dell Latitude 5430,IT-LAP-0001,IT Equipment,School Asset,'
        'Example laptop,YPJ Kuala Kencana,Computer Lab A,active,'
        '1250.00,2023-08-15,4,2026-08-15,DL5430X01,Dell Technologies,Maria Hernandez'
    )
    buf = io.BytesIO(f"{header}\n{example}\n".encode())
    headers = {"Content-Disposition": 'attachment; filename="asset-import-template.csv"'}
    return StreamingResponse(buf, media_type="text/csv", headers=headers)


@api.post("/assets/import")
async def bulk_import_assets(
    file: UploadFile = File(...),
    dry_run: bool = Query(True),
    user: dict = Depends(admin_required),
):
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 5 MB)")

    rows = _parse_import_file(content, file.filename or "upload.csv")
    if not rows:
        raise HTTPException(400, "File is empty or has no data rows")

    # --- validate ---
    conn = get_conn()
    existing_tags = {
        r[0] for r in conn.execute("SELECT asset_tag FROM assets").fetchall()
    }

    seen_tags: set = set()
    preview = []
    for i, row in enumerate(rows, start=2):  # row 1 = header
        errs = _validate_import_row(row, i, seen_tags)
        tag = (row.get("asset_tag") or "").strip()
        if tag and tag in existing_tags:
            errs.append(f'asset_tag "{tag}" already exists in the database')
        if tag:
            seen_tags.add(tag)
        preview.append({
            "row": i,
            "name": (row.get("name") or "").strip(),
            "asset_tag": tag,
            "category": (row.get("category") or "").strip(),
            "campus": (row.get("campus") or "").strip(),
            "location": (row.get("location") or "").strip(),
            "status": (row.get("status") or "active").strip(),
            "errors": errs,
        })

    total   = len(preview)
    invalid = sum(1 for p in preview if p["errors"])
    valid   = total - invalid

    if dry_run:
        conn.close()
        return {"dry_run": True, "total": total, "valid": valid, "invalid": invalid, "preview": preview}

    # --- actual import ---
    ts = now_iso()
    imported = 0
    skipped  = 0
    import_errors = []

    for i, row in enumerate(rows, start=2):
        p = preview[i - 2]
        if p["errors"]:
            skipped += 1
            import_errors.append({"row": i, "errors": p["errors"]})
            continue
        try:
            aid = str(uuid.uuid4())
            name      = (row.get("name") or "").strip()
            tag       = (row.get("asset_tag") or "").strip()
            cat       = (row.get("category") or "").strip()
            atype     = (row.get("asset_type") or "School Asset").strip()
            desc      = (row.get("description") or "").strip()
            campus    = (row.get("campus") or "").strip()
            loc       = (row.get("location") or "").strip()
            status    = (row.get("status") or "active").strip()
            price_s   = (row.get("purchase_price") or "0").strip()
            price     = float(price_s) if price_s else 0.0
            pdate     = (row.get("purchase_date") or "").strip() or None
            life_s    = (row.get("useful_life_years") or "5").strip()
            life      = int(life_s) if life_s else 5
            wdate     = (row.get("warranty_end_date") or "").strip() or None
            serial    = (row.get("serial_number") or "").strip()
            supplier  = (row.get("supplier") or "").strip()
            owner     = (row.get("assigned_to_name") or "").strip()

            conn.execute("""
                INSERT INTO assets
                  (id,name,asset_tag,category,asset_type,description,campus,location,status,
                   purchase_price,purchase_date,useful_life_years,warranty_end_date,
                   serial_number,supplier,assigned_to_user_id,assigned_to_name,
                   photo_path,documents,created_at,updated_at,created_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                aid, name, tag, cat, atype, desc, campus, loc, status,
                price, pdate, life, wdate, serial, supplier,
                None, owner or None,
                None, "[]", ts, ts, user["id"],
            ))
            if owner:
                conn.execute("""
                    INSERT INTO ownership_logs
                      (id,asset_id,from_name,from_user_id,to_name,to_user_id,note,by,by_name,at)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                """, (str(uuid.uuid4()), aid, None, None, owner, None,
                      "Bulk import", user["id"], user["name"], ts))
            imported += 1
        except Exception as exc:
            skipped += 1
            import_errors.append({"row": i, "errors": [str(exc)]})

    conn.commit()
    conn.close()
    return {"dry_run": False, "imported": imported, "skipped": skipped, "errors": import_errors}


@api.get("/assets/{aid}")
async def get_asset(aid: str, user: dict = Depends(get_current_user)):
    conn = get_conn()
    row = conn.execute("SELECT * FROM assets WHERE id = ?", (aid,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, "Asset not found")
    asset = row_to_dict(row)
    asset["depreciation"] = calc_depreciation(asset["purchase_price"], asset["purchase_date"], asset.get("useful_life_years", 5))
    asset["ownership_history"] = [
        dict(r) for r in conn.execute("SELECT * FROM ownership_logs WHERE asset_id = ? ORDER BY at DESC", (aid,)).fetchall()
    ]
    asset["faults"] = [
        dict(r) for r in conn.execute("SELECT * FROM faults WHERE asset_id = ? ORDER BY created_at DESC", (aid,)).fetchall()
    ]
    asset["compliance"] = [
        dict(r) for r in conn.execute("SELECT * FROM compliance WHERE asset_id = ? ORDER BY due_date ASC", (aid,)).fetchall()
    ]
    conn.close()
    return asset


@api.put("/assets/{aid}")
async def update_asset(aid: str, body: AssetUpdate, user: dict = Depends(get_current_user)):
    updates = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    if not updates:
        raise HTTPException(400, "No fields to update")
    updates["updated_at"] = now_iso()
    if "documents" in updates:
        updates["documents"] = json.dumps(updates["documents"])
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    params = list(updates.values()) + [aid]
    conn = get_conn()
    res = conn.execute(f"UPDATE assets SET {set_clause} WHERE id = ?", params)
    if res.rowcount == 0:
        conn.close()
        raise HTTPException(404, "Asset not found")
    conn.commit()
    row = conn.execute("SELECT * FROM assets WHERE id = ?", (aid,)).fetchone()
    conn.close()
    return row_to_dict(row)


@api.delete("/assets/{aid}")
async def delete_asset(aid: str, _: dict = Depends(admin_required)):
    conn = get_conn()
    conn.execute("DELETE FROM assets WHERE id = ?", (aid,))
    conn.execute("DELETE FROM ownership_logs WHERE asset_id = ?", (aid,))
    conn.execute("DELETE FROM faults WHERE asset_id = ?", (aid,))
    conn.commit()
    conn.close()
    return {"ok": True}


@api.post("/assets/{aid}/transfer")
async def transfer_asset(aid: str, body: TransferIn, user: dict = Depends(get_current_user)):
    conn = get_conn()
    row = conn.execute("SELECT * FROM assets WHERE id = ?", (aid,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, "Asset not found")
    asset = row_to_dict(row)
    log_id = str(uuid.uuid4())
    ts = now_iso()
    conn.execute("""
        INSERT INTO ownership_logs (id,asset_id,from_name,from_user_id,to_name,to_user_id,note,by,by_name,at)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (log_id, aid, asset.get("assigned_to_name"), asset.get("assigned_to_user_id"),
          body.new_owner_name, body.new_owner_user_id, body.note or "", user["id"], user["name"], ts))
    conn.execute(
        "UPDATE assets SET assigned_to_name=?, assigned_to_user_id=?, updated_at=? WHERE id=?",
        (body.new_owner_name, body.new_owner_user_id, ts, aid)
    )
    conn.commit()
    log = dict(conn.execute("SELECT * FROM ownership_logs WHERE id = ?", (log_id,)).fetchone())
    conn.close()
    return log


# ---- Faults ----
@api.get("/faults")
async def list_faults(user: dict = Depends(get_current_user), status: Optional[str] = None):
    conn = get_conn()
    if status:
        rows = conn.execute("SELECT * FROM faults WHERE status = ? ORDER BY created_at DESC", (status,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM faults ORDER BY created_at DESC").fetchall()
    faults = [dict(r) for r in rows]
    for f in faults:
        a = conn.execute("SELECT name, asset_tag FROM assets WHERE id = ?", (f["asset_id"],)).fetchone()
        f["asset_name"] = a["name"] if a else "(deleted)"
        f["asset_tag"] = a["asset_tag"] if a else ""
    conn.close()
    return faults


@api.post("/faults")
async def create_fault(body: FaultIn, user: dict = Depends(get_current_user)):
    conn = get_conn()
    a = conn.execute("SELECT id FROM assets WHERE id = ?", (body.asset_id,)).fetchone()
    if not a:
        conn.close()
        raise HTTPException(404, "Asset not found")
    fid = str(uuid.uuid4())
    ts = now_iso()
    conn.execute("""
        INSERT INTO faults (id,asset_id,title,description,severity,status,
            reported_by,reported_by_name,resolution_note,resolved_at,created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (fid, body.asset_id, body.title, body.description, body.severity, "open",
          user["id"], user["name"], None, None, ts))
    conn.commit()
    row = conn.execute("SELECT * FROM faults WHERE id = ?", (fid,)).fetchone()
    conn.close()
    return dict(row)


@api.put("/faults/{fid}")
async def update_fault(fid: str, body: FaultUpdate, user: dict = Depends(get_current_user)):
    updates = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    if updates.get("status") == "resolved":
        updates["resolved_at"] = now_iso()
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    params = list(updates.values()) + [fid]
    conn = get_conn()
    res = conn.execute(f"UPDATE faults SET {set_clause} WHERE id = ?", params)
    if res.rowcount == 0:
        conn.close()
        raise HTTPException(404, "Fault not found")
    conn.commit()
    row = conn.execute("SELECT * FROM faults WHERE id = ?", (fid,)).fetchone()
    conn.close()
    return dict(row)


# ---- Compliance ----
@api.get("/compliance")
async def list_compliance(user: dict = Depends(get_current_user)):
    conn = get_conn()
    rows = conn.execute("SELECT * FROM compliance ORDER BY due_date ASC").fetchall()
    items = [dict(r) for r in rows]
    today = datetime.now(timezone.utc)
    for c in items:
        try:
            due = datetime.fromisoformat(c["due_date"].replace("Z", "+00:00"))
            if due.tzinfo is None:
                due = due.replace(tzinfo=timezone.utc)
            if c.get("status") == "pending" and due < today:
                c["status"] = "overdue"
        except Exception:
            pass
        if c.get("asset_id"):
            a = conn.execute("SELECT name FROM assets WHERE id = ?", (c["asset_id"],)).fetchone()
            c["asset_name"] = a["name"] if a else None
    conn.close()
    return items


@api.post("/compliance")
async def create_compliance(body: ComplianceIn, _: dict = Depends(admin_required)):
    conn = get_conn()
    cid = str(uuid.uuid4())
    ts = now_iso()
    conn.execute("""
        INSERT INTO compliance (id,asset_id,title,description,category,due_date,frequency,
            status,completed_date,note,created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (cid, body.asset_id, body.title, body.description or "", body.category,
          body.due_date, body.frequency, "pending", None, None, ts))
    conn.commit()
    row = conn.execute("SELECT * FROM compliance WHERE id = ?", (cid,)).fetchone()
    conn.close()
    return dict(row)


@api.put("/compliance/{cid}")
async def update_compliance(cid: str, body: ComplianceUpdate, _: dict = Depends(admin_required)):
    updates = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    if updates.get("status") == "completed" and not updates.get("completed_date"):
        updates["completed_date"] = now_iso()
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    params = list(updates.values()) + [cid]
    conn = get_conn()
    res = conn.execute(f"UPDATE compliance SET {set_clause} WHERE id = ?", params)
    if res.rowcount == 0:
        conn.close()
        raise HTTPException(404, "Compliance item not found")
    conn.commit()
    row = conn.execute("SELECT * FROM compliance WHERE id = ?", (cid,)).fetchone()
    conn.close()
    return dict(row)


@api.delete("/compliance/{cid}")
async def delete_compliance(cid: str, _: dict = Depends(admin_required)):
    conn = get_conn()
    conn.execute("DELETE FROM compliance WHERE id = ?", (cid,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ---- Activity log ----
@api.get("/activity")
async def get_activity(
    _: dict = Depends(get_current_user),
    page: int = Query(1, ge=1),
    limit: int = Query(30, ge=1, le=100),
    event_type: Optional[str] = None,
):
    conn = get_conn()
    offset = (page - 1) * limit

    type_filter = ""
    if event_type and event_type != "all":
        type_filter = f"WHERE type = '{event_type}'"

    rows = conn.execute(f"""
        SELECT * FROM (
            SELECT
                'asset_created' AS type,
                a.id AS ref_id,
                a.name AS subject,
                u.name AS actor,
                a.category AS detail,
                a.campus AS campus,
                a.created_at AS at
            FROM assets a
            LEFT JOIN users u ON u.id = a.created_by

            UNION ALL

            SELECT
                'ownership_transfer' AS type,
                ol.asset_id AS ref_id,
                a.name AS subject,
                ol.by_name AS actor,
                ol.to_name AS detail,
                a.campus AS campus,
                ol.at AS at
            FROM ownership_logs ol
            LEFT JOIN assets a ON a.id = ol.asset_id

            UNION ALL

            SELECT
                'fault_reported' AS type,
                f.asset_id AS ref_id,
                a.name AS subject,
                f.reported_by_name AS actor,
                f.title AS detail,
                a.campus AS campus,
                f.created_at AS at
            FROM faults f
            LEFT JOIN assets a ON a.id = f.asset_id

            UNION ALL

            SELECT
                'fault_resolved' AS type,
                f.asset_id AS ref_id,
                a.name AS subject,
                f.reported_by_name AS actor,
                f.title AS detail,
                a.campus AS campus,
                f.resolved_at AS at
            FROM faults f
            LEFT JOIN assets a ON a.id = f.asset_id
            WHERE f.resolved_at IS NOT NULL

            UNION ALL

            SELECT
                'compliance_added' AS type,
                c.id AS ref_id,
                c.title AS subject,
                NULL AS actor,
                c.category AS detail,
                NULL AS campus,
                c.created_at AS at
            FROM compliance c

            UNION ALL

            SELECT
                'user_created' AS type,
                u.id AS ref_id,
                u.name AS subject,
                NULL AS actor,
                u.role AS detail,
                u.campus AS campus,
                u.created_at AS at
            FROM users u
        ) events
        {type_filter}
        ORDER BY at DESC
        LIMIT ? OFFSET ?
    """, (limit, offset)).fetchall()

    total = conn.execute(f"""
        SELECT COUNT(*) FROM (
            SELECT created_at AS at, 'asset_created' AS type FROM assets
            UNION ALL SELECT at, 'ownership_transfer' FROM ownership_logs
            UNION ALL SELECT created_at, 'fault_reported' FROM faults
            UNION ALL SELECT resolved_at, 'fault_resolved' FROM faults WHERE resolved_at IS NOT NULL
            UNION ALL SELECT created_at, 'compliance_added' FROM compliance
            UNION ALL SELECT created_at, 'user_created' FROM users
        ) e {type_filter}
    """).fetchone()[0]

    conn.close()
    return {"items": [dict(r) for r in rows], "total": total, "page": page, "limit": limit}


# ---- Dashboard ----
@api.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(get_current_user)):
    conn = get_conn()
    rows = conn.execute("SELECT * FROM assets").fetchall()
    assets = [row_to_dict(r) for r in rows]
    total_assets = len(assets)
    total_value = sum(a.get("purchase_price", 0) for a in assets)
    current_value = 0
    by_category: dict = {}
    by_status: dict = {}
    for a in assets:
        d = calc_depreciation(a["purchase_price"], a["purchase_date"], a.get("useful_life_years", 5))
        current_value += d["current_value"]
        by_category[a["category"]] = by_category.get(a["category"], 0) + 1
        by_status[a["status"]] = by_status.get(a["status"], 0) + 1

    today = datetime.now(timezone.utc)
    soon = today + timedelta(days=60)
    warranty_alerts = []
    for a in assets:
        we = a.get("warranty_end_date")
        if not we:
            continue
        try:
            d = datetime.fromisoformat(we.replace("Z", "+00:00"))
            if d.tzinfo is None:
                d = d.replace(tzinfo=timezone.utc)
            if d <= soon:
                warranty_alerts.append({
                    "id": a["id"], "name": a["name"], "asset_tag": a["asset_tag"],
                    "warranty_end_date": we, "expired": d < today,
                })
        except Exception:
            pass

    open_faults = conn.execute(
        "SELECT COUNT(*) FROM faults WHERE status IN ('open','in_progress')"
    ).fetchone()[0]
    recent_fault_rows = conn.execute("SELECT * FROM faults ORDER BY created_at DESC LIMIT 5").fetchall()
    recent_faults = []
    for f in recent_fault_rows:
        fd = dict(f)
        a = conn.execute("SELECT name FROM assets WHERE id = ?", (fd["asset_id"],)).fetchone()
        fd["asset_name"] = a["name"] if a else "(deleted)"
        recent_faults.append(fd)

    compliance_rows = conn.execute("SELECT * FROM compliance").fetchall()
    overdue = 0
    upcoming = 0
    for c in compliance_rows:
        try:
            d = datetime.fromisoformat(c["due_date"].replace("Z", "+00:00"))
            if d.tzinfo is None:
                d = d.replace(tzinfo=timezone.utc)
            if c["status"] == "completed":
                continue
            if d < today:
                overdue += 1
            elif d <= today + timedelta(days=30):
                upcoming += 1
        except Exception:
            pass

    timeline = []
    if assets:
        for yr_offset in range(0, 6):
            ts = today + timedelta(days=365 * yr_offset)
            total = 0
            for a in assets:
                try:
                    pdate = datetime.fromisoformat(a["purchase_date"].replace("Z", "+00:00"))
                    if pdate.tzinfo is None:
                        pdate = pdate.replace(tzinfo=timezone.utc)
                    age = (ts - pdate).days / 365.25
                    annual = a["purchase_price"] / max(a.get("useful_life_years", 5), 1)
                    accumulated = min(annual * age, a["purchase_price"])
                    val = max(a["purchase_price"] - accumulated, 0)
                    total += val
                except Exception:
                    pass
            timeline.append({"year": ts.year, "value": round(total, 2)})

    conn.close()
    return {
        "total_assets": total_assets,
        "total_purchase_value": round(total_value, 2),
        "current_book_value": round(current_value, 2),
        "depreciated_value": round(total_value - current_value, 2),
        "by_category": by_category,
        "by_status": by_status,
        "warranty_alerts": warranty_alerts,
        "open_faults": open_faults,
        "recent_faults": recent_faults,
        "compliance_overdue": overdue,
        "compliance_upcoming": upcoming,
        "depreciation_timeline": timeline,
    }


# ---- Transfer requests ----
class TransferRequestIn(BaseModel):
    new_owner_name: str
    note: Optional[str] = ""


@api.post("/transfer-requests")
async def create_transfer_request(asset_id: str, body: TransferRequestIn, user: dict = Depends(get_current_user)):
    conn = get_conn()
    asset = conn.execute("SELECT id, name, asset_tag, assigned_to_name FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if not asset:
        conn.close()
        raise HTTPException(404, "Asset not found")
    rid = str(uuid.uuid4())
    conn.execute("""
        INSERT INTO transfer_requests
            (id, asset_id, asset_name, asset_tag, from_name, to_name, note,
             requested_by, requested_by_name, status, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,'pending',?)
    """, (rid, asset_id, asset["name"], asset["asset_tag"], asset["assigned_to_name"],
          body.new_owner_name, body.note or "", user["id"], user["name"], now_iso()))
    conn.commit()
    row = conn.execute("SELECT * FROM transfer_requests WHERE id = ?", (rid,)).fetchone()
    conn.close()
    return dict(row)


@api.get("/transfer-requests")
async def list_transfer_requests(user: dict = Depends(get_current_user), status: Optional[str] = None):
    conn = get_conn()
    if user["role"] == "admin":
        base = "SELECT * FROM transfer_requests"
        rows = conn.execute(base + (" WHERE status = ?" if status else "") + " ORDER BY created_at DESC",
                            ([status] if status else [])).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM transfer_requests WHERE requested_by = ? ORDER BY created_at DESC", (user["id"],)
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@api.get("/transfer-requests/pending-count")
async def pending_count(_: dict = Depends(admin_required)):
    conn = get_conn()
    n = conn.execute("SELECT COUNT(*) FROM transfer_requests WHERE status = 'pending'").fetchone()[0]
    conn.close()
    return {"count": n}


@api.put("/transfer-requests/{rid}/approve")
async def approve_transfer(rid: str, user: dict = Depends(admin_required)):
    conn = get_conn()
    req = conn.execute("SELECT * FROM transfer_requests WHERE id = ?", (rid,)).fetchone()
    if not req:
        conn.close()
        raise HTTPException(404, "Request not found")
    req = dict(req)
    if req["status"] != "pending":
        conn.close()
        raise HTTPException(400, "Request is not pending")
    ts = now_iso()
    # Perform the actual ownership transfer
    log_id = str(uuid.uuid4())
    conn.execute("""
        INSERT INTO ownership_logs (id,asset_id,from_name,from_user_id,to_name,to_user_id,note,by,by_name,at)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (log_id, req["asset_id"], req["from_name"], None, req["to_name"], None,
          req["note"], user["id"], user["name"], ts))
    conn.execute("UPDATE assets SET assigned_to_name=?, updated_at=? WHERE id=?",
                 (req["to_name"], ts, req["asset_id"]))
    conn.execute("""
        UPDATE transfer_requests SET status='approved', reviewed_by=?, reviewed_by_name=?, reviewed_at=? WHERE id=?
    """, (user["id"], user["name"], ts, rid))
    conn.commit()
    conn.close()
    return {"ok": True}


@api.put("/transfer-requests/{rid}/reject")
async def reject_transfer(rid: str, reason: Optional[str] = None, user: dict = Depends(admin_required)):
    conn = get_conn()
    req = conn.execute("SELECT * FROM transfer_requests WHERE id = ?", (rid,)).fetchone()
    if not req:
        conn.close()
        raise HTTPException(404, "Request not found")
    if dict(req)["status"] != "pending":
        conn.close()
        raise HTTPException(400, "Request is not pending")
    ts = now_iso()
    conn.execute("""
        UPDATE transfer_requests SET status='rejected', reviewed_by=?, reviewed_by_name=?, reviewed_at=?, reject_reason=? WHERE id=?
    """, (user["id"], user["name"], ts, reason or "", rid))
    conn.commit()
    conn.close()
    return {"ok": True}


def _xlsx_response(wb, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


def _style_header(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
        cell.alignment = Alignment(horizontal="center")


@api.get("/reports/asset-inventory")
async def report_asset_inventory(_: dict = Depends(admin_required)):
    import openpyxl
    conn = get_conn()
    rows = conn.execute("""
        SELECT name, asset_tag, category, campus, location, status,
               assigned_to_name, purchase_price, purchase_date, useful_life_years,
               warranty_end_date, serial_number, supplier, created_at
        FROM assets ORDER BY name
    """).fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Inventory"
    headers = ["Asset Name", "Tag", "Category", "Campus", "Room/Location", "Status",
               "Assigned To", "Purchase Price (USD)", "Purchase Date", "Useful Life (yrs)",
               "Warranty End", "Serial Number", "Supplier", "Created At"]
    ws.append(headers)
    _style_header(ws)
    for r in rows:
        d = dict(r)
        dep = calc_depreciation(d["purchase_price"], d["purchase_date"], d.get("useful_life_years", 5))
        ws.append([
            d["name"], d["asset_tag"], d["category"], d["campus"], d["location"],
            d["status"], d["assigned_to_name"] or "", d["purchase_price"],
            d["purchase_date"][:10] if d["purchase_date"] else "",
            d["useful_life_years"], d["warranty_end_date"][:10] if d["warranty_end_date"] else "",
            d["serial_number"] or "", d["supplier"] or "", d["created_at"][:10],
        ])
    today = datetime.now().strftime("%Y%m%d")
    return _xlsx_response(wb, f"asset-inventory-{today}.xlsx")


@api.get("/reports/depreciation")
async def report_depreciation(_: dict = Depends(admin_required)):
    import openpyxl
    conn = get_conn()
    rows = conn.execute("""
        SELECT name, asset_tag, category, campus, purchase_price, purchase_date, useful_life_years
        FROM assets ORDER BY name
    """).fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Depreciation Schedule"
    headers = ["Asset Name", "Tag", "Category", "Campus", "Purchase Price (USD)",
               "Purchase Date", "Useful Life (yrs)", "Annual Depreciation (USD)",
               "Accumulated Depreciation (USD)", "Current Book Value (USD)", "Age (yrs)"]
    ws.append(headers)
    _style_header(ws)
    for r in rows:
        d = dict(r)
        dep = calc_depreciation(d["purchase_price"], d["purchase_date"], d.get("useful_life_years", 5))
        ws.append([
            d["name"], d["asset_tag"], d["category"], d["campus"],
            d["purchase_price"], d["purchase_date"][:10] if d["purchase_date"] else "",
            d["useful_life_years"],
            round(dep["annual"], 2), round(dep["accumulated"], 2),
            round(dep["current_value"], 2), round(dep["age_years"], 2),
        ])
    today = datetime.now().strftime("%Y%m%d")
    return _xlsx_response(wb, f"depreciation-schedule-{today}.xlsx")


@api.get("/reports/faults")
async def report_faults(_: dict = Depends(admin_required)):
    import openpyxl
    conn = get_conn()
    rows = conn.execute("""
        SELECT f.title, f.description, f.severity, f.status,
               a.name AS asset_name, a.asset_tag, a.campus,
               f.reported_by_name, f.resolution_note, f.resolved_at, f.created_at
        FROM faults f
        LEFT JOIN assets a ON a.id = f.asset_id
        ORDER BY f.created_at DESC
    """).fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fault Report"
    headers = ["Fault Title", "Description", "Severity", "Status", "Asset Name",
               "Asset Tag", "Campus", "Reported By", "Resolution Note", "Resolved At", "Reported At"]
    ws.append(headers)
    _style_header(ws)
    for r in rows:
        d = dict(r)
        ws.append([
            d["title"], d["description"], d["severity"], d["status"],
            d["asset_name"] or "", d["asset_tag"] or "", d["campus"] or "",
            d["reported_by_name"],
            d["resolution_note"] or "", d["resolved_at"][:10] if d["resolved_at"] else "",
            d["created_at"][:10],
        ])
    today = datetime.now().strftime("%Y%m%d")
    return _xlsx_response(wb, f"fault-report-{today}.xlsx")


@api.get("/reports/compliance")
async def report_compliance(_: dict = Depends(admin_required)):
    import openpyxl
    conn = get_conn()
    rows = conn.execute("""
        SELECT c.title, c.description, c.category, c.frequency, c.due_date,
               c.status, c.completed_date, c.note,
               a.name AS asset_name, a.asset_tag, c.created_at
        FROM compliance c
        LEFT JOIN assets a ON a.id = c.asset_id
        ORDER BY c.due_date
    """).fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Register"
    headers = ["Title", "Description", "Category", "Frequency", "Due Date", "Status",
               "Completed Date", "Note", "Asset Name", "Asset Tag", "Created At"]
    ws.append(headers)
    _style_header(ws)
    for r in rows:
        d = dict(r)
        ws.append([
            d["title"], d["description"] or "", d["category"], d["frequency"],
            d["due_date"][:10] if d["due_date"] else "", d["status"],
            d["completed_date"][:10] if d["completed_date"] else "",
            d["note"] or "", d["asset_name"] or "", d["asset_tag"] or "",
            d["created_at"][:10],
        ])
    today = datetime.now().strftime("%Y%m%d")
    return _xlsx_response(wb, f"compliance-register-{today}.xlsx")


@api.get("/reports/ownership")
async def report_ownership(_: dict = Depends(admin_required)):
    import openpyxl
    conn = get_conn()
    rows = conn.execute("""
        SELECT a.name AS asset_name, a.asset_tag, a.campus, a.category,
               ol.from_name, ol.to_name, ol.note, ol.by_name, ol.at
        FROM ownership_logs ol
        LEFT JOIN assets a ON a.id = ol.asset_id
        ORDER BY ol.at DESC
    """).fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ownership History"
    headers = ["Asset Name", "Asset Tag", "Campus", "Category",
               "Transferred From", "Transferred To", "Note", "Actioned By", "Date"]
    ws.append(headers)
    _style_header(ws)
    for r in rows:
        d = dict(r)
        ws.append([
            d["asset_name"] or "", d["asset_tag"] or "", d["campus"] or "", d["category"] or "",
            d["from_name"] or "", d["to_name"] or "", d["note"] or "",
            d["by_name"], d["at"][:10] if d["at"] else "",
        ])
    today = datetime.now().strftime("%Y%m%d")
    return _xlsx_response(wb, f"ownership-history-{today}.xlsx")


@api.get("/backup")
async def backup_db(user: dict = Depends(admin_required)):
    if not DB_PATH.exists():
        raise HTTPException(404, "Database file not found")
    return FileResponse(
        str(DB_PATH),
        media_type="application/octet-stream",
        filename=f"ypj-backup-{datetime.now().strftime('%Y%m%d')}.db",
    )


# ---- Loan requests ----

class LoanRequestIn(BaseModel):
    purpose: Optional[str] = ""
    start_date: Optional[str] = None
    end_date: Optional[str] = None


@api.post("/assets/{aid}/loan-request")
async def create_loan_request(aid: str, body: LoanRequestIn, user: dict = Depends(get_current_user)):
    conn = get_conn()
    asset = conn.execute("SELECT id,name,asset_tag FROM assets WHERE id = ?", (aid,)).fetchone()
    if not asset:
        conn.close()
        raise HTTPException(404, "Asset not found")
    rid = str(uuid.uuid4())
    ts = now_iso()
    conn.execute(
        """INSERT INTO loan_requests
           (id,asset_id,asset_name,asset_tag,requested_by,requested_by_name,purpose,start_date,end_date,status,created_at)
           VALUES (?,?,?,?,?,?,?,?,?,'pending',?)""",
        (rid, aid, asset["name"], asset["asset_tag"], user["id"], user["name"],
         body.purpose or "", body.start_date, body.end_date, ts),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM loan_requests WHERE id = ?", (rid,)).fetchone()
    req = dict(row)

    # Email: confirmation to borrower, CC supervisor + campus heads + notify inbox
    subject = f'[YPJ Assets] Loan request submitted: {req["asset_name"]}'
    body_html = _loan_html(req, "Loan Request Submitted", [
        ("Asset",      f'{req["asset_name"]} ({req["asset_tag"]})'),
        ("Requested by", req["requested_by_name"]),
        ("Purpose",    req.get("purpose") or "—"),
        ("From",       req.get("start_date") or "—"),
        ("Until",      req.get("end_date") or "—"),
        ("Status",     "Pending approval"),
    ], "<p>Your request has been submitted and is awaiting admin approval.</p>")
    cc = get_loan_cc(req, conn, extra=[LOAN_NOTIFY_EMAIL] if LOAN_NOTIFY_EMAIL else [])
    send_email(user["email"], subject, body_html, cc=cc)

    conn.close()
    return req


@api.get("/loan-requests")
async def list_loan_requests(user: dict = Depends(get_current_user), status: Optional[str] = None):
    conn = get_conn()
    if user["role"] == "admin":
        if status and status != "all":
            rows = conn.execute(
                "SELECT * FROM loan_requests WHERE status=? ORDER BY created_at DESC", (status,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM loan_requests ORDER BY created_at DESC"
            ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM loan_requests WHERE requested_by=? ORDER BY created_at DESC", (user["id"],)
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@api.get("/loan-requests/pending-count")
async def loan_pending_count(_: dict = Depends(admin_required)):
    conn = get_conn()
    n = conn.execute("SELECT COUNT(*) FROM loan_requests WHERE status='pending'").fetchone()[0]
    conn.close()
    return {"count": n}


@api.put("/loan-requests/{rid}/approve")
async def approve_loan(rid: str, user: dict = Depends(admin_required)):
    conn = get_conn()
    req = conn.execute("SELECT * FROM loan_requests WHERE id=?", (rid,)).fetchone()
    if not req:
        conn.close()
        raise HTTPException(404, "Loan request not found")
    now = now_iso()
    conn.execute(
        "UPDATE loan_requests SET status='approved',reviewed_by=?,reviewed_by_name=?,reviewed_at=? WHERE id=?",
        (user["id"], user["name"], now, rid),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM loan_requests WHERE id=?", (rid,)).fetchone()
    req = dict(row)

    # Email borrower; CC supervisor + campus heads
    borrower = conn.execute("SELECT email FROM users WHERE id=?", (req["requested_by"],)).fetchone()
    if borrower:
        subject = f'[YPJ Assets] Loan approved: {req["asset_name"]}'
        body_html = _loan_html(req, "Loan Request Approved", [
            ("Asset",       f'{req["asset_name"]} ({req["asset_tag"]})'),
            ("Borrower",    req["requested_by_name"]),
            ("Approved by", req["reviewed_by_name"]),
            ("From",        req.get("start_date") or "—"),
            ("Until",       req.get("end_date") or "—"),
        ], "<p>Your loan request has been approved. The asset will be delivered to you soon.</p>")
        cc = get_loan_cc(req, conn)
        send_email(borrower["email"], subject, body_html, cc=cc)

    conn.close()
    return req


@api.put("/loan-requests/{rid}/reject")
async def reject_loan(rid: str, reason: Optional[str] = None, user: dict = Depends(admin_required)):
    conn = get_conn()
    req = conn.execute("SELECT * FROM loan_requests WHERE id=?", (rid,)).fetchone()
    if not req:
        conn.close()
        raise HTTPException(404, "Loan request not found")
    conn.execute(
        "UPDATE loan_requests SET status='rejected',reviewed_by=?,reviewed_by_name=?,reviewed_at=?,reject_reason=? WHERE id=?",
        (user["id"], user["name"], now_iso(), reason or "", rid),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM loan_requests WHERE id=?", (rid,)).fetchone()
    req = dict(row)

    # Email borrower; CC supervisor only (campus heads not needed for rejections)
    borrower = conn.execute("SELECT email FROM users WHERE id=?", (req["requested_by"],)).fetchone()
    if borrower:
        subject = f'[YPJ Assets] Loan request not approved: {req["asset_name"]}'
        reason_line = req.get("reject_reason") or ""
        detail = [
            ("Asset",       f'{req["asset_name"]} ({req["asset_tag"]})'),
            ("Rejected by", req["reviewed_by_name"]),
        ]
        if reason_line:
            detail.append(("Reason", reason_line))
        body_html = _loan_html(req, "Loan Request Not Approved", detail,
                               "<p>Please contact the admin if you have questions.</p>")
        u = conn.execute("SELECT supervisor FROM users WHERE id=?", (req["requested_by"],)).fetchone()
        cc = [u["supervisor"]] if u and u["supervisor"] else []
        send_email(borrower["email"], subject, body_html, cc=cc)

    conn.close()
    return req


@api.put("/loan-requests/{rid}/deliver")
async def deliver_loan(
    rid: str,
    condition: str = Form(...),
    note: str = Form(""),
    image: Optional[UploadFile] = File(None),
    user: dict = Depends(admin_required),
):
    """Record pre-check condition and optionally upload an image, then mark as active."""
    conn = get_conn()
    req = conn.execute("SELECT * FROM loan_requests WHERE id=?", (rid,)).fetchone()
    if not req:
        conn.close()
        raise HTTPException(404, "Loan request not found")
    if dict(req)["status"] != "approved":
        conn.close()
        raise HTTPException(400, "Loan must be in 'approved' status to deliver")

    img_path = ""
    if image and image.filename:
        ext = image.filename.rsplit(".", 1)[-1].lower() if "." in image.filename else "bin"
        file_id = str(uuid.uuid4())
        rel_path = f"loans/{rid}/pre_{file_id}.{ext}"
        dest = UPLOADS_DIR / "loans" / rid
        dest.mkdir(parents=True, exist_ok=True)
        data = await image.read()
        (dest / f"pre_{file_id}.{ext}").write_bytes(data)
        img_path = rel_path
        conn.execute(
            "INSERT INTO files (id,storage_path,original_filename,content_type,size,uploaded_by,is_deleted,created_at) VALUES (?,?,?,?,?,?,0,?)",
            (file_id, rel_path, image.filename, image.content_type or "application/octet-stream", len(data), user["id"], now_iso()),
        )

    conn.execute(
        """UPDATE loan_requests
           SET status='active', pre_condition=?, pre_condition_note=?, pre_image_path=?, delivered_at=?
           WHERE id=?""",
        (condition, note, img_path, now_iso(), rid),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM loan_requests WHERE id=?", (rid,)).fetchone()
    req = dict(row)

    # Email borrower; CC supervisor + campus heads
    borrower = conn.execute("SELECT email FROM users WHERE id=?", (req["requested_by"],)).fetchone()
    if borrower:
        cond_labels = {"good": "Good", "trouble": "Has issues", "broken": "Broken", "missing": "Missing"}
        subject = f'[YPJ Assets] Asset delivered: {req["asset_name"]}'
        body_html = _loan_html(req, "Asset Delivered", [
            ("Asset",       f'{req["asset_name"]} ({req["asset_tag"]})'),
            ("Borrower",    req["requested_by_name"]),
            ("Condition",   cond_labels.get(condition, condition)),
            ("Note",        note or "—"),
            ("Return by",   req.get("end_date") or "—"),
        ], "<p>The asset is now in your care. Please return it by the agreed date in the same or better condition.</p>")
        cc = get_loan_cc(req, conn)
        send_email(borrower["email"], subject, body_html, cc=cc)

    conn.close()
    return req


@api.put("/loan-requests/{rid}/return")
async def return_loan(
    rid: str,
    condition: str = Form(...),
    note: str = Form(""),
    image: Optional[UploadFile] = File(None),
    user: dict = Depends(admin_required),
):
    """Record post-check condition and optionally upload an image, then mark as returned."""
    conn = get_conn()
    req = conn.execute("SELECT * FROM loan_requests WHERE id=?", (rid,)).fetchone()
    if not req:
        conn.close()
        raise HTTPException(404, "Loan request not found")
    if dict(req)["status"] != "active":
        conn.close()
        raise HTTPException(400, "Loan must be in 'active' status to return")

    img_path = ""
    if image and image.filename:
        ext = image.filename.rsplit(".", 1)[-1].lower() if "." in image.filename else "bin"
        file_id = str(uuid.uuid4())
        rel_path = f"loans/{rid}/post_{file_id}.{ext}"
        dest = UPLOADS_DIR / "loans" / rid
        dest.mkdir(parents=True, exist_ok=True)
        data = await image.read()
        (dest / f"post_{file_id}.{ext}").write_bytes(data)
        img_path = rel_path
        conn.execute(
            "INSERT INTO files (id,storage_path,original_filename,content_type,size,uploaded_by,is_deleted,created_at) VALUES (?,?,?,?,?,?,0,?)",
            (file_id, rel_path, image.filename, image.content_type or "application/octet-stream", len(data), user["id"], now_iso()),
        )

    conn.execute(
        """UPDATE loan_requests
           SET status='returned', post_condition=?, post_condition_note=?, post_image_path=?, returned_at=?
           WHERE id=?""",
        (condition, note, img_path, now_iso(), rid),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM loan_requests WHERE id=?", (rid,)).fetchone()
    req = dict(row)

    # Email borrower (confirmation) + CC supervisor + campus heads + loan notify inbox
    borrower = conn.execute("SELECT email FROM users WHERE id=?", (req["requested_by"],)).fetchone()
    if borrower:
        cond_labels = {"good": "Good", "trouble": "Has issues", "broken": "Broken", "missing": "Missing"}
        subject = f'[YPJ Assets] Return recorded: {req["asset_name"]}'
        body_html = _loan_html(req, "Asset Return Recorded", [
            ("Asset",            f'{req["asset_name"]} ({req["asset_tag"]})'),
            ("Borrower",         req["requested_by_name"]),
            ("Return condition", cond_labels.get(condition, condition)),
            ("Note",             note or "—"),
            ("Returned at",      req.get("returned_at") or "—"),
        ], "<p>Thank you for returning the asset. The loan is now closed.</p>")
        cc = get_loan_cc(req, conn, extra=[LOAN_NOTIFY_EMAIL] if LOAN_NOTIFY_EMAIL else [])
        send_email(borrower["email"], subject, body_html, cc=cc)

    conn.close()
    return req


# ---- Mount ----
app.include_router(api)
app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")

# Serve React build in production (frontend/build lives one level up from backend/)
BUILD_DIR = ROOT_DIR.parent / "frontend" / "build"
if BUILD_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(BUILD_DIR / "static")), name="react-static")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_spa(full_path: str):
        index = BUILD_DIR / "index.html"
        return HTMLResponse(index.read_text())

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', 'http://localhost:3000,http://localhost:8000').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---- Startup: init DB + seed ----
def seed_data():
    conn = get_conn()
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@school.edu")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    staff_email = os.environ.get("STAFF_EMAIL", "staff@school.edu")
    staff_password = os.environ.get("STAFF_PASSWORD", "staff123")

    # Find the admin by email OR by role (handles email changes)
    admin = conn.execute("SELECT id FROM users WHERE email = ?", (admin_email,)).fetchone()
    if not admin:
        existing_admin = conn.execute("SELECT id, email FROM users WHERE role = 'admin' LIMIT 1").fetchone()
        if existing_admin:
            # Migrate email to new value from env var
            conn.execute("UPDATE users SET email = ? WHERE id = ?", (admin_email, existing_admin["id"]))
            admin_id = existing_admin["id"]
            logger.info(f"Updated admin email to: {admin_email}")
        else:
            admin_id = str(uuid.uuid4())
            conn.execute(
                "INSERT INTO users (id,email,name,role,department,password_hash,created_at) VALUES (?,?,?,?,?,?,?)",
                (admin_id, admin_email, "Admin User", "admin", "Administration", hash_password(admin_password), now_iso()),
            )
            logger.info(f"Seeded admin: {admin_email}")
    else:
        admin_id = admin["id"]

    # Remove demo staff user if still present
    demo_staff = conn.execute("SELECT id FROM users WHERE email = ?", (staff_email,)).fetchone()
    if demo_staff:
        conn.execute("DELETE FROM users WHERE id = ?", (demo_staff["id"],))
        logger.info(f"Removed demo staff user: {staff_email}")
    staff_id = admin_id  # fallback so demo asset seeding still works

    asset_count = conn.execute("SELECT COUNT(*) FROM assets").fetchone()[0]
    if asset_count == 0:
        ts = now_iso()
        # (name, tag, category, desc, campus, location, status, price, pdate, life, warranty, serial, supplier, assignee_id, assignee_name)
        demo = [
            ("Dell Latitude 5430 Laptop", "IT-LAP-0001", "IT Equipment",
             "Teacher laptop with Windows 11 Pro", "YPJ Kuala Kencana", "Computer Lab A", "active",
             1250.00, "2023-08-15", 4, "2026-08-15", "DL5430X01", "Dell Technologies", staff_id, "Maria Hernandez"),
            ("Epson PowerLite Projector", "AV-PRJ-0014", "AV Equipment",
             "Classroom projector 3500 lumens", "YPJ Tembagapura", "Room 204", "active",
             680.00, "2022-03-10", 6, "2025-03-10", "EPL3500-204", "Epson", None, "Room 204"),
            ("Microscope Set Olympus CX23", "LAB-MIC-0007", "Lab Equipment",
             "Set of 8 student microscopes", "YPJ Kuala Kencana", "Biology Lab", "active",
             4200.00, "2021-09-01", 10, "2024-09-01", "OLY-CX23-S8", "Olympus", staff_id, "Maria Hernandez"),
            ("Steelcase Student Desks (40)", "FUR-DSK-0040", "Furniture",
             "Set of 40 height-adjustable desks", "YPJ Kuala Kencana", "Room 112", "active",
             8000.00, "2020-07-20", 12, "2030-07-20", "SC-LOT-112", "Steelcase", None, "Room 112"),
            ("Toyota Hiace School Bus", "VEH-BUS-0001", "Vehicles",
             "16-seater school transport van", "YPJ Tembagapura", "Main Garage", "active",
             35000.00, "2019-04-12", 8, "2024-04-12", "TYT-HIACE-19", "Toyota", None, "Transport Dept"),
            ("iPad Pro 11\" Cart (30 units)", "IT-IPD-0030", "IT Equipment",
             "Mobile cart of iPads for classroom use", "YPJ Kuala Kencana", "Library", "active",
             27000.00, "2024-01-20", 5, "2027-01-20", "IPD-CART-LIB1", "Apple", None, "Library"),
            ("Smart Interactive Whiteboard", "AV-IWB-0003", "AV Equipment",
             "Smart Board 6065 65-inch", "YPJ Tembagapura", "Room 301", "in_repair",
             3500.00, "2022-11-05", 7, "2025-11-05", "SB6065-301", "Smart Technologies", None, "Room 301"),
            ("Casio Digital Piano CDP-S110", "MUS-PIA-0002", "Music Equipment",
             "Compact digital piano for music room", "YPJ Kuala Kencana", "Music Room", "active",
             540.00, "2023-02-14", 8, "2026-02-14", "CDS110-002", "Casio", None, "Music Room"),
        ]
        asset_ids = []
        for row in demo:
            aid = str(uuid.uuid4())
            asset_ids.append(aid)
            (name, tag, cat, desc, campus, loc, status, price, pdate, life, warranty, serial, supplier, assignee_id, assignee_name) = row
            conn.execute("""
                INSERT INTO assets (id,name,asset_tag,category,description,campus,location,status,
                    purchase_price,purchase_date,useful_life_years,warranty_end_date,
                    serial_number,supplier,assigned_to_user_id,assigned_to_name,
                    photo_path,documents,created_at,updated_at,created_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (aid, name, tag, cat, desc, campus, loc, status, price, pdate, life, warranty,
                  serial, supplier, assignee_id, assignee_name, None, "[]", ts, ts, admin_id))
            conn.execute("""
                INSERT INTO ownership_logs (id,asset_id,from_name,from_user_id,to_name,to_user_id,note,by,by_name,at)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (str(uuid.uuid4()), aid, None, None, assignee_name, assignee_id, "Initial assignment", admin_id, "Admin User", ts))

        # Sample faults (whiteboard = index 6, projector = index 1)
        conn.execute("""
            INSERT INTO faults (id,asset_id,title,description,severity,status,reported_by,reported_by_name,resolution_note,resolved_at,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (str(uuid.uuid4()), asset_ids[6], "Whiteboard touch sensors not responding",
              "Bottom-left quadrant unresponsive to touch input.", "high", "in_progress",
              staff_id, "Maria Hernandez", None, None, ts))
        conn.execute("""
            INSERT INTO faults (id,asset_id,title,description,severity,status,reported_by,reported_by_name,resolution_note,resolved_at,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (str(uuid.uuid4()), asset_ids[1], "Projector lamp dim",
              "Image is noticeably dim, lamp may need replacement.", "medium", "open",
              staff_id, "Maria Hernandez", None, None, ts))

        # Sample compliance
        soon_due = (datetime.now(timezone.utc) + timedelta(days=14)).isoformat()
        overdue_due = (datetime.now(timezone.utc) - timedelta(days=10)).isoformat()
        future_due = (datetime.now(timezone.utc) + timedelta(days=120)).isoformat()
        conn.execute("""
            INSERT INTO compliance (id,asset_id,title,description,category,due_date,frequency,status,completed_date,note,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (str(uuid.uuid4()), asset_ids[4], "Annual School Bus Safety Inspection",
              "Mandatory DOT-equivalent inspection", "Safety", soon_due, "annual", "pending", None, None, ts))
        conn.execute("""
            INSERT INTO compliance (id,asset_id,title,description,category,due_date,frequency,status,completed_date,note,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (str(uuid.uuid4()), None, "Fire Extinguisher Recertification",
              "All campus fire extinguishers to be recertified", "Safety", overdue_due, "annual", "pending", None, None, ts))
        conn.execute("""
            INSERT INTO compliance (id,asset_id,title,description,category,due_date,frequency,status,completed_date,note,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (str(uuid.uuid4()), asset_ids[2], "Lab Equipment Calibration",
              "Annual calibration of microscopes", "Inspection", future_due, "annual", "pending", None, None, ts))

        logger.info("Seeded demo assets, faults, and compliance items")

    conn.commit()
    conn.close()


@app.on_event("startup")
async def on_startup():
    init_db()
    seed_data()
    # Start background scheduler for daily return reminders
    scheduler = BackgroundScheduler(timezone="Asia/Jakarta")
    scheduler.add_job(send_return_reminders, "cron", hour=8, minute=0)
    scheduler.start()
    mode = "production" if BUILD_DIR.exists() else "development"
    logger.info(f"Startup complete [{mode}] — DB: {DB_PATH}")
